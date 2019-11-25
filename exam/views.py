from django.shortcuts import render, render_to_response
import sqlite3
import json
import openpyxl
import os
import pandas as pd
from django.core import serializers
from django.http import HttpResponse, JsonResponse
from exam.models import result
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone
from django.db import transaction
from datetime import datetime
import logging
from django.contrib.auth.models import Permission, User
from django.shortcuts import get_object_or_404




logger = logging.getLogger(__name__)



def download(request):
    try:
        user_name = request.user.username
        conn = sqlite3.connect("first.db")
        cur = conn.cursor()
        sql = """
                select example_id || " " || b.content, a.select_id || " " || c.name, a.title_id || " " || d.name  ,date
                from exam_result a, TB_EXAM b, TB_TITLE c, TB_TITLE d
                where a.example_id = b.id
                and check_YN = 'N'
                and a.select_id = c.id
                and a.title_id = d.id
                and a.user_name = ?
                group by title_id, example_id
                order by a.date desc, a.title_id, a.example_id
            """
        cur.execute(sql, (user_name,))
        rows = cur.fetchall()

        #wb = openpyxl.load_workbook()
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = '틀린문항'

        sheet.cell(row = 1, column = 1, value = '문제')
        sheet.cell(row = 1, column = 2, value = '선택한답')
        sheet.cell(row = 1, column = 3, value = '정답')
        sheet.cell(row = 1, column = 4, value = '날짜')
        
        for i in range(len(rows)):
            for j in range(len(rows[i])):
                sheet.cell(row = i+2, column = j+1, value = rows[i][j])           

        wb.save('result.xlsx')
        wb.close()
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        filepath = os.path.join(BASE_DIR, 'result.xlsx')
        filename = os.path.basename(filepath)

        with open(filepath, 'rb') as f:
            response = HttpResponse(f, content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)

    finally:
        conn.close()
        

    return response

def statics_false(request):
    try:
        if request.user.has_perm('test'):
            print("has perm")
        user_name = request.user.username
        conn = sqlite3.connect("first.db")
        cur = conn.cursor()
        sql = """
                select example_id || " " || b.content, a.select_id || " " || c.name, a.title_id || " " || d.name  ,date
                from exam_result a, TB_EXAM b, TB_TITLE c, TB_TITLE d
                where a.example_id = b.id
                and check_YN = 'N'
                and a.select_id = c.id
                and a.title_id = d.id
                and a.user_name = ?
                group by title_id, example_id
                order by a.date desc, a.title_id, a.example_id
            """
        cur.execute(sql, (user_name,))
        statics_list = cur.fetchall()
        size = 10
        paginator = Paginator(statics_list,size)
        page = request.GET.get('page')
        posts = paginator.get_page(page)
        context = {
            "posts" : posts,
        }
    finally:
        conn.close()
    
    return render(request, "statics_false.html", context)

def data(request):
    try:
        user_name = request.user.username
        conn = sqlite3.connect("first.db")
        cur = conn.cursor()

        query = cur.execute("select * from exam_result")
        cols = [column[0] for column in query.description]
        df = pd.DataFrame.from_records(data=query.fetchall(), columns=cols)

        df['_date'] = df['date'].str.slice(0,10)
        con_right = (df['check_yn'] == 'Y') & (df['user_name'] == user_name)
        con_all = df['user_name'] == user_name

        grouped_right = df.loc[con_right,"check_yn"].groupby(df["_date"]).count()
        grouped_all = df.loc[con_all,"check_yn"].groupby(df["_date"]).count()
        right_rate = grouped_right / grouped_all 
        right_rate = pd.DataFrame({'dates':right_rate.index, 'rates':right_rate.values}).fillna(0)
        right_rate['rates'] = right_rate['rates'] * 100
        dates = []
        rates = []
        dates.append('날짜')
        dates = dates + right_rate['dates'].values.tolist()

        rates.append('정답률')
        rates = rates + right_rate['rates'].values.tolist()

        data = {
            'columns':[
                dates,
                rates,
            ]
        }
        

        

        # sql = """
        #         select date, round(false*1.0/total*100)
        #         from (select date, count(date) as false, (select count(a.date) from exam_result a where a.date = b.date) as total
        #         from exam_result b
        #         where user_name = ?
        #         and check_YN = 'Y'
        #         group by date)               
        #     """
        # cur.execute(sql, (user_name,))
        # rows = cur.fetchall()
        # dates = ['일자', ]
        # rates = ['정답률',]
        # for row in rows:
        #     dates.append("'"+row[0][2:])
        #     rates.append(row[1])
        # data = {
        #     'columns':[
        #         dates,
        #         rates,
        #     ]
        # }
    finally:
        conn.close()

    return HttpResponse(json.dumps(data),content_type="text/json")

def chart_rate(request):
    return render(request, "chart_rate.html")




def statics(request):    
    user_name = request.user.username
    conn = sqlite3.connect("first.db")
    cur = conn.cursor()
    sql = """
            select title_id || " " || name, count(*) as cnt 
            from (select a.title_id, b.name from exam_result a, TB_TITLE b where a.user_name = ? and a.title_id=b.id and a.check_yn = 'N') 
            group by title_id order by cnt DESC
        """
    cur.execute(sql, (user_name,))
    statics_list = cur.fetchall()
    size = 10
    paginator = Paginator(statics_list,size)
    page = request.GET.get('page')
    posts = paginator.get_page(page)
    #statics = json.dumps(statics_list)
    context = {
        #"test" : statics_list,
        "posts" : posts,
        #"statics" : statics,
    }
    return render(request, "statics.html", context)
    

def getExample(request):
    try: 
        conn = sqlite3.connect("first.db")
        cur = conn.cursor()
        sql = """
                SELECT id, content 
                FROM TB_EXAM 
                WHERE id not in (
                    select distinct example_id 
                    from exam_result
                    where session_id = ?
                    and check_yn = 'Y'
                    )
                ORDER BY RANDOM() LIMIT 1
              """
        cur.execute(sql, (request.session.session_key,))
        row = cur.fetchone()
        example_id = row[0]
        title_id = row[0][0:5]
        #length = len(title_id)
        if title_id[4] == ".": 
            title_id = row[0][0:6]

        exam_content = row[1]
        sql = "SELECT name, content FROM TB_TITLE WHERE id = ?"
        cur.execute(sql, (title_id,))
        row = cur.fetchone()
        title_name = row[0]
        title_content = row[1]

        #cur.execute("SELECT content FROM TB_TITLE WHERE id = ?",(title_id))
                    
        context = {
            "title_id": title_id,
            "title_name": title_name,
            "title_content": title_content,
            "example_id": example_id,
            "example_content": exam_content,
        }

    finally:
        conn.close()

    return render(request, "exam.html", context)


def getTitleContent(request):
    try:
        title_id = request.GET["title_id"]
        # b = request.user.has_perm('test')
        # c = request.user.has_perm('aaa')
        conn = sqlite3.connect("first.db")
        cur = conn.cursor()
        cur.execute("SELECT name, content FROM TB_TITLE WHERE id = ?",(title_id,))
        row = cur.fetchone()
        name = row[0]
        content = row[1]
        return JsonResponse({
            "false_title_name" : name,
            "false_title_content" : content,
        }, json_dumps_params = {'ensure_ascii': True})
    #except:

    finally:
        conn.close()



def sendResult(request):
    #try:
    select_id = request.GET["select_id"]
    example_id = request.GET["example_id"]
    session_id = request.session.session_key
    user_name = request.user.username
    #date = request.user.last_login.strftime('%Y%m%d')
    #date = datetime.now().strftime('%Y%m%d')
    date = datetime.now()

    if example_id[5] != ".": 
        title_id = example_id[0:6]
    else:
        title_id = example_id[0:5]

    if select_id == title_id :
        check_yn = 'Y'
    else :
        check_yn = 'N'
    try:
        conn = sqlite3.connect("first.db")
        cur = conn.cursor()
        sql = """
            insert into exam_result(select_id, title_id, example_id, session_id, user_name, date, check_yn)
            values(?,?,?,?,?,?,?)
        """

        cur.execute(sql,(select_id, title_id, example_id, session_id, user_name, date, check_yn))
        conn.commit()
    except:
       conn.rollback()
       print("error")

    #result(select_id=select_id, title_id = title_id, example_id=example_id, session_id=session_id, user_name=user_name, date=date, check_yn=check_yn).save()
    finally:
        conn.close()

    return JsonResponse({
    }, json_dumps_params = {'ensure_ascii': True})
    # finally:
    #     conn.close()